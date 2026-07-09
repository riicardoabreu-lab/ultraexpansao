"""
Gera relatório de provedores (caixas e clientes) por bairro, a partir
dos CSVs colocados em csv-input/, no formato exportado do Map Marker
(colunas: Folder name, Folder color, Latitude, Longitude, Title, ...).

Convenção usada (mesma do fluxo manual da Infolink):
- Folder name segue o padrão MUNICIPIO/BAIRRO/CAIXAS/PROVEDOR[/TIPO]
- Cada linha do CSV = uma caixa (CTO)
- QTD. CAIXAS  = contagem de linhas por provedor
- QTD. CLIENTES = soma da coluna "Title" por provedor

Roda automaticamente pelo GitHub Actions (.github/workflows/gerar-relatorio.yml)
sempre que um novo CSV é adicionado/alterado em csv-input/.
"""

import csv
import glob
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "csv-input")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "relatorios")

VIOLET = "4C1D95"
CYAN = "06B6D4"
LIGHT_GRAY = "F3F4F6"
FONT_NAME = "Arial"


def parse_csv(path):
    with open(path, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        return None

    by_provider = defaultdict(lambda: {"caixas": 0, "clientes": 0})
    municipio, bairro = None, None

    for r in rows:
        folder = r.get("Folder name", "")
        parts = [p.strip() for p in folder.split("/")]
        if len(parts) < 4:
            continue

        if municipio is None:
            municipio, bairro = parts[0], parts[1]

        provider = parts[3]
        if len(parts) > 4 and parts[4]:
            provider = f"{provider} {parts[4]}"

        title = (r.get("Title") or "").strip()
        qtd_clientes = int(title) if title.isdigit() else 0

        by_provider[provider]["caixas"] += 1
        by_provider[provider]["clientes"] += qtd_clientes

    return {
        "municipio": municipio,
        "bairro": bairro,
        "providers": dict(sorted(by_provider.items())),
    }


def build_report(data, output_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Relatorio"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells("A1:C1")
    sheet["A1"] = f"{data['bairro']} — {data['municipio']}"
    sheet["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", start_color=VIOLET, end_color=VIOLET)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    headers = ["PROVEDOR", "QTD. CAIXAS", "QTD. CLIENTES"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col, value=header)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=CYAN, end_color=CYAN)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row_idx = 3
    for provider, vals in data["providers"].items():
        sheet.cell(row=row_idx, column=1, value=provider).border = border
        sheet.cell(row=row_idx, column=2, value=vals["caixas"]).border = border
        sheet.cell(row=row_idx, column=3, value=vals["clientes"]).border = border
        if row_idx % 2 == 0:
            for col in range(1, 4):
                sheet.cell(row=row_idx, column=col).fill = PatternFill(
                    "solid", start_color=LIGHT_GRAY, end_color=LIGHT_GRAY
                )
        row_idx += 1

    total_row = row_idx
    first_data_row = 3
    last_data_row = row_idx - 1
    sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(
        name=FONT_NAME, bold=True
    )
    sheet.cell(
        row=total_row, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})"
    ).font = Font(name=FONT_NAME, bold=True)
    sheet.cell(
        row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})"
    ).font = Font(name=FONT_NAME, bold=True)
    for col in range(1, 4):
        sheet.cell(row=total_row, column=col).fill = PatternFill(
            "solid", start_color="E5E7EB", end_color="E5E7EB"
        )
        sheet.cell(row=total_row, column=col).border = border

    for col, width in zip("ABC", (32, 16, 16)):
        sheet.column_dimensions[col].width = width

    for row in sheet.iter_rows(min_row=3, max_row=total_row):
        for cell in row:
            if cell.column > 1:
                cell.alignment = Alignment(horizontal="center")

    wb.save(output_path)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    csv_files = glob.glob(os.path.join(INPUT_DIR, "*.csv")) + glob.glob(
        os.path.join(INPUT_DIR, "*.CSV")
    )

    if not csv_files:
        print("Nenhum CSV encontrado em csv-input/.")
        return

    gerados = []
    for path in sorted(set(csv_files)):
        data = parse_csv(path)
        if not data or not data["bairro"]:
            print(f"Aviso: não foi possível interpretar {path}, pulando.")
            continue

        out_name = f"{data['bairro'].upper()}.xlsx"
        out_path = os.path.join(OUTPUT_DIR, out_name)
        build_report(data, out_path)
        gerados.append(out_name)
        print(f"Gerado: relatorios/{out_name}")

    print(f"\n{len(gerados)} relatório(s) gerado(s): {', '.join(gerados)}")


if __name__ == "__main__":
    main()
